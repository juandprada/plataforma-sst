"""Convierte EMPRESAS.xlsx -> data/empresas.json.

La primera fila del Excel son los encabezados (EMPRESA, LOGO, NIT, ...).
Cada fila siguiente es una empresa. Las claves se normalizan a MAYUSCULAS con
guion bajo (p.ej. "REPRESENTANTE LEGAL" -> "REPRESENTANTE_LEGAL") para usarse
como tokens {{TOKEN}} en las plantillas.

Uso:
    python tools/xlsx_to_json.py \
        --xlsx "../1.PLANEAR/Input/EMPRESAS.xlsx" \
        --out data/empresas.json
"""
from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


def normalize_key(raw: str) -> str:
    """'Representante Legal' -> 'REPRESENTANTE_LEGAL'."""
    s = str(raw).strip().upper()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^A-Z0-9_]", "", s)
    return s


def slugify(value: str) -> str:
    """'ARMAR ALIMENTOS SAS' -> 'armar-alimentos-sas' (id estable/URL-safe)."""
    s = unicodedata.normalize("NFKD", str(value))
    s = s.encode("ascii", "ignore").decode("ascii").lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--xlsx",
        default="../1.PLANEAR/Input/EMPRESAS.xlsx",
        help="Ruta al Excel de empresas.",
    )
    ap.add_argument("--out", default="data/empresas.json")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("El Excel no tiene filas.")

    header = [normalize_key(c) if c is not None else "" for c in rows[0]]

    empresas: list[dict] = []
    for row in rows[1:]:
        # Salta filas totalmente vacias.
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        data: dict[str, str] = {}
        for key, cell in zip(header, row):
            if not key:
                continue
            data[key] = "" if cell is None else str(cell).strip()
        nombre = data.get("EMPRESA", "").strip()
        if not nombre:
            continue
        data["_id"] = slugify(nombre)
        empresas.append(data)

    empresas.sort(key=lambda e: e.get("EMPRESA", ""))

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(empresas, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"OK: {len(empresas)} empresas -> {out_path}")
    if empresas:
        print("Tokens disponibles:", ", ".join(k for k in empresas[0] if k != "_id"))


if __name__ == "__main__":
    main()
