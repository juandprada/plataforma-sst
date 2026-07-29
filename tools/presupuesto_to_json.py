"""Genera data/presupuesto.json a partir de 1.PLANEAR/3.PRESUPUESTO.xlsx.

El .xlsx es una calculadora: la tabla del presupuesto (columnas A-D) sale de 5
parámetros por empresa (columnas J-Q, una por empresa) mediante fórmulas:

    B3 = Pago                      -> Asesoría Externa (valor unitario mensual), cant 12
    B4 = 20000 * trabajadores      -> Elementos de Oficina
    B5 = 30000 * trabajadores      -> Capacitaciones en SG-SST
    C8 = extintores                -> Mantenimiento y recarga de extintores
    C9 = botiquines                -> Implementos para el botiquín
    C12 = trabajadores             -> Elementos de Protección Personal
    C19 = aires acondicionados     -> Mantenimiento de Aires Acondicionados
    C22 = trabajadores             -> Exámenes Periódicos
    B23 = 30000 * trabajadores     -> Semana de la Salud
    D25 = suma de subtotales,  D26 = D25 / 12

Este script extrae SOLO los parámetros por empresa (lo que cambia); el catálogo de
ítems y precios vive en el JSON para poder actualizarlo por inflación sin tocar código.
Uso:  python tools/presupuesto_to_json.py
"""

import json
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO.parent / "1.PLANEAR" / "3.PRESUPUESTO.xlsx"
SALIDA = REPO / "data" / "presupuesto.json"

# Columna del .xlsx -> _id de empresa en data/empresas.json. Las columnas que no
# corresponden a ninguna empresa registrada (PREMIUM, POSEM, cauchos, AREPAS) se ignoran.
COLUMNA_A_EMPRESA = {
    "ck": "ck-comercializadora-un-mundo-de-oportunidades-sas",
    "gaf": "gaf-technology-solutions-sas",
    "MEDALLAS": "medalleria-deportiva-del-valle",
    "tr": "te-recuperamos-sas",
}

# Filas de parámetros dentro del bloque de la derecha del .xlsx.
FILA_PARAM = {
    2: "trabajadores",
    3: "extintores",
    4: "botiquines",
    5: "aires_acondicionados",
    6: "pago_mensual_asesoria",
}

# Catálogo: ítems del presupuesto agrupados como en el .docx. Cada ítem define
# cómo se obtienen su valor unitario y su cantidad:
#   vu   -> {"fijo": n}  |  {"fijo": n, "por": "<param>"}  |  {"param": "<param>"}
#   cant -> {"fijo": n}  |  {"param": "<param>"}
CATALOGO = [
    {
        "grupo": "Sistema de Gestión de Seguridad y Salud en el Trabajo",
        "items": [
            {"nombre": "Asesoría Externa SG-SST",
             "vu": {"param": "pago_mensual_asesoria"}, "cant": {"fijo": 12}},
            {"nombre": "Elementos de Oficina",
             "vu": {"fijo": 20000, "por": "trabajadores"}, "cant": {"fijo": 1}},
            {"nombre": "Capacitaciones en SG-SST",
             "vu": {"fijo": 30000, "por": "trabajadores"}, "cant": {"fijo": 1}},
        ],
    },
    {
        "grupo": "Brigadas de Emergencias",
        "items": [
            {"nombre": "Mantenimiento y recarga de extintores",
             "vu": {"fijo": 50000}, "cant": {"param": "extintores"}},
            {"nombre": "Implementos para el botiquín",
             "vu": {"fijo": 30000}, "cant": {"param": "botiquines"}},
        ],
    },
    {
        "grupo": "Seguridad Industrial",
        "items": [
            {"nombre": "Elementos de Protección Personal",
             "vu": {"fijo": 100000}, "cant": {"param": "trabajadores"}},
            {"nombre": "Fumigación", "vu": {"fijo": 200000}, "cant": {"fijo": 1}},
        ],
    },
    {
        "grupo": "Higiene Industrial",
        "items": [
            {"nombre": "Desinfectantes", "vu": {"fijo": 50000}, "cant": {"fijo": 1}},
            {"nombre": "Caja de Toallas Desechables", "vu": {"fijo": 50000}, "cant": {"fijo": 1}},
            {"nombre": "Caja de Jabones", "vu": {"fijo": 50000}, "cant": {"fijo": 1}},
            {"nombre": "Mantenimiento de Aires Acondicionados",
             "vu": {"fijo": 100000}, "cant": {"param": "aires_acondicionados"}},
        ],
    },
    {
        "grupo": "Salud Laboral",
        "items": [
            {"nombre": "Exámenes Periódicos",
             "vu": {"fijo": 25000}, "cant": {"param": "trabajadores"}},
            {"nombre": "Semana de la Salud",
             "vu": {"fijo": 30000, "por": "trabajadores"}, "cant": {"fijo": 1}},
        ],
    },
]

# Parámetros de respaldo para empresas que no están en el .xlsx (columna G del Excel).
DEFAULT = {
    "trabajadores": 4,
    "extintores": 1,
    "botiquines": 1,
    "aires_acondicionados": 0,
    "pago_mensual_asesoria": 250000,
}


def leer_parametros():
    """Lee el bloque de parámetros por empresa (columnas J-Q) del .xlsx."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Sheet1"]
    por_empresa = {}
    for col in range(10, 18):  # J..Q
        etiqueta = ws.cell(row=1, column=col).value
        empresa_id = COLUMNA_A_EMPRESA.get(str(etiqueta).strip() if etiqueta else "")
        if not empresa_id:
            continue
        datos = {}
        for fila, clave in FILA_PARAM.items():
            v = ws.cell(row=fila, column=col).value
            # El "Pago" suele estar vacío por empresa: se hereda del default.
            datos[clave] = int(v) if isinstance(v, (int, float)) else DEFAULT[clave]
        por_empresa[empresa_id] = datos
    return por_empresa


def main():
    por_empresa = leer_parametros()
    # Los ajustes manuales (valor/cantidad por ítem, del panel de la app) NO salen del
    # .xlsx: se conservan tal cual si ya estaban en el JSON.
    ajustes = {}
    if SALIDA.exists():
        previo = json.loads(SALIDA.read_text(encoding="utf-8"))
        ajustes = previo.get("ajustes_por_empresa", {})
    salida = {
        "_origen": "1.PLANEAR/3.PRESUPUESTO.xlsx (regenerar con tools/presupuesto_to_json.py)",
        "moneda": "COP",
        "parametros_default": DEFAULT,
        "parametros_por_empresa": por_empresa,
        "_ajustes_por_empresa": (
            'Overrides permanentes de valor unitario/cantidad por ítem: '
            '{"<empresa_id>": {"<grupo>|<item>": {"vu": n, "cant": n}}}. '
            "Se pegan con el botón 'Copiar ajustes' del panel de la app."
        ),
        "ajustes_por_empresa": ajustes,
        "catalogo": CATALOGO,
    }
    SALIDA.parent.mkdir(exist_ok=True)
    SALIDA.write_text(json.dumps(salida, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"OK -> {SALIDA.relative_to(REPO)}  ({len(por_empresa)} empresas con datos propios)")
    for eid, p in por_empresa.items():
        print(f"   {eid:46} {p}")


if __name__ == "__main__":
    main()
